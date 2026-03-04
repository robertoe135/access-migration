# app.py
# Streamlit app: Fixture Schedule (MF) -> TEMPLATE.xlsx format
# - Upload your TEMPLATE.xlsx (expects sheet: tbeFixtureTypeDetails)
# - Upload your MF fixture schedule CSV (the "format you originally imported")
# - Exports a CSV matching the template headers (150 cols) with your rules:
#   1) If "VOID" appears anywhere in a fixture block -> set template column "VOID" = True
#   2) Unit type: if unit is ln.ft -> "ln.ft" else "each"
#   3) Dim protocol: if schedule says 0-10V -> "0-10v (TVI)"
#   4) Pull Location + Protection + LumenOutput + InputLoad per fixture
#   5) Do NOT modify CatalogNo text (keeps notes like "*TBC by architect" etc.)
#   6) Any other not-apparent fields copy from the template's first data row

import io
import re
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
import streamlit as st
import openpyxl


# -----------------------------
# Helpers
# -----------------------------

FIXTURE_CODE_RE = re.compile(r"^[A-Z]{1,3}(?:-\d+)?$")

def _s(x) -> str:
    if x is None:
        return ""
    if isinstance(x, float) and np.isnan(x):
        return ""
    return str(x)

def is_fixture_code(x) -> bool:
    return bool(FIXTURE_CODE_RE.fullmatch(_s(x).strip()))

def block_contains_void(block: pd.DataFrame) -> bool:
    # Check for the word VOID anywhere, case-insensitive, whole-word-ish
    # We’ll treat "VOID" appearing inside longer strings as a match too.
    for col in block.columns:
        ser = block[col]
        for v in ser.values:
            t = _s(v)
            if t and re.search(r"\bVOID\b", t, flags=re.IGNORECASE):
                return True
    return False

def pick_labeled_value(block: pd.DataFrame, label: str,
                       label_col: str = "Unnamed: 3", value_col: str = "Unnamed: 4") -> str:
    for _, r in block.iterrows():
        lab = _s(r.get(label_col, "")).strip()
        if lab.upper().startswith(label.upper()):
            return _s(r.get(value_col, ""))
    return ""

def pick_lumens(block: pd.DataFrame) -> str:
    # In the MF schedule CSV, lumens appear in Unnamed: 7 (seen as numeric)
    col = "Unnamed: 7"
    if col not in block.columns:
        return ""
    for v in block[col].values:
        t = _s(v).strip()
        if not t:
            continue
        if re.fullmatch(r"\d+(\.\d+)?", t):
            return t
    # fallback: first non-empty
    for v in block[col].values:
        t = _s(v).strip()
        if t:
            return t
    return ""

def pick_unit(block: pd.DataFrame) -> str:
    # In the MF schedule CSV, the unit for lumens row is often in Unnamed: 8 (e.g., "ea.")
    # We normalize to:
    #  - "ln.ft" if source unit indicates linear feet
    #  - otherwise "each"
    unit_cols = ["Unnamed: 8", "Unnamed: 6", "Unnamed: 9"]
    raw = ""
    for c in unit_cols:
        if c in block.columns:
            for v in block[c].values:
                t = _s(v).strip()
                if t:
                    raw = t
                    break
        if raw:
            break

    raw_low = raw.lower()
    if "ln.ft" in raw_low or "ln/ft" in raw_low or "linear" in raw_low:
        return "ln.ft"
    return "each"

def pick_input_load(block: pd.DataFrame) -> str:
    # Find a token like "19 W" anywhere in the block
    for _, r in block.iterrows():
        for c in block.columns:
            t = _s(r.get(c, "")).strip()
            if not t:
                continue
            m = re.search(r"\b\d+(\.\d+)?\s*[Ww]\b", t)
            if m:
                token = m.group(0)
                # normalize lowercase w -> W, keep spacing as seen
                token = re.sub(r"w\b", "W", token)
                return token
    return ""

def pick_catalog_scored_exact(block: pd.DataFrame) -> str:
    """
    Pick the most catalog-like string from Unnamed: 1 while preserving text EXACTLY.
    (Keeps suffix notes like '*TBC by architect', extra asterisks, etc.)
    """
    if "Unnamed: 1" not in block.columns:
        return ""

    candidates: List[Tuple[int, int, str]] = []
    for k in range(1, len(block)):
        val = block.iloc[k].get("Unnamed: 1", "")
        if pd.isna(val):
            continue
        txt = str(val)  # exact
        t = txt.strip()
        if not t:
            continue
        low = t.lower()

        # Skip obvious non-catalog lines
        if "http" in low or "www" in low or "@" in t or "#ref" in low:
            continue
        if re.search(r"\(\d{3}\)", t) or re.search(r"\d{3}[-\s]\d{3}[-\s]\d{4}", t):
            continue

        score = 0
        if "_" in t: score += 10
        if re.search(r"\d", t): score += 3
        if "*" in t: score += 2
        if len(t) > 18: score += 2
        if re.search(r"[A-Z]{2,}\d", t): score += 2

        # Penalize "pretty name" like "Fraxion 3" (no underscores)
        if re.fullmatch(r"[A-Za-z]+\s*\d+", t) and "_" not in t:
            score -= 6

        candidates.append((score, k, txt))

    if not candidates:
        return ""

    # Highest score, then longer string
    candidates.sort(key=lambda x: (-x[0], -len(x[2])))
    return candidates[0][2]

def find_best_unit_field(headers: List[str]) -> Optional[str]:
    # Try to locate the template column that should receive "ln.ft" / "each"
    priority = ["UnitType", "Unit", "Units", "QtyUnit", "QuantityUnit"]
    for p in priority:
        if p in headers:
            return p
    # fallback: any header containing 'unit' but not 'units' (unless that's all we have)
    unitish = [h for h in headers if h and re.search(r"unit", str(h), re.IGNORECASE)]
    if not unitish:
        return None
    # prefer ones that look like a "type" field
    for h in unitish:
        if re.search(r"type", str(h), re.IGNORECASE):
            return h
    return unitish[0]

def load_template_defaults(template_bytes: bytes) -> Tuple[List[str], Dict[str, object]]:
    wb = openpyxl.load_workbook(io.BytesIO(template_bytes), data_only=True)
    if "tbeFixtureTypeDetails" not in wb.sheetnames:
        raise ValueError("Template must contain a sheet named 'tbeFixtureTypeDetails'.")
    ws = wb["tbeFixtureTypeDetails"]

    headers = [c.value for c in ws[1]]
    defaults = [c.value for c in ws[2]]
    if not headers or any(h is None for h in headers):
        raise ValueError("Template header row (row 1) appears to be missing/invalid.")
    default_row = dict(zip(headers, defaults))
    return headers, default_row

def transform(schedule_df: pd.DataFrame, headers: List[str], default_row: Dict[str, object]) -> pd.DataFrame:
    # Determine fixture blocks
    first_col = schedule_df.columns[0]
    fixture_idx = [i for i, v in enumerate(schedule_df[first_col]) if is_fixture_code(v)]
    if not fixture_idx:
        raise ValueError("No fixture designations found (e.g., MA, MA-1, MB...).")

    unit_field = find_best_unit_field(headers)

    out_rows: List[Dict[str, object]] = []

    for j, start in enumerate(fixture_idx):
        end = fixture_idx[j + 1] if j + 1 < len(fixture_idx) else len(schedule_df)
        block = schedule_df.iloc[start:end]

        fixture_type = _s(block.iloc[0][first_col]).strip()
        manufacturer = _s(block.iloc[0].get("Unnamed: 1", "")).strip()
        base_desc = _s(block.iloc[0].get("Unnamed: 3", "")).strip()
        dim_proto = _s(block.iloc[0].get("Unnamed: 17", "")).strip()

        if dim_proto.upper() == "0-10V":
            dim_proto = "0-10v (TVI)"

        catalog = pick_catalog_scored_exact(block)  # exact text
        protection = pick_labeled_value(block, "PROTECTION:").strip()
        location = pick_labeled_value(block, "LOCATION:").strip()
        lumens = pick_lumens(block).strip()
        input_load = pick_input_load(block).strip()
        unit_value = pick_unit(block)

        void_flag = block_contains_void(block)

        out = dict(default_row)  # copy template row 2 defaults
        # required mappings
        if "Type" in out:
            out["Type"] = fixture_type

        if "Manufacturer" in out and manufacturer:
            out["Manufacturer"] = manufacturer

        if "CatalogNo" in out and catalog:
            out["CatalogNo"] = catalog

        if "BaseDescription" in out and base_desc:
            out["BaseDescription"] = base_desc

        if "Protection" in out and protection:
            out["Protection"] = protection

        if "Location" in out and location:
            out["Location"] = location

        if "DimProtocol" in out and dim_proto:
            out["DimProtocol"] = dim_proto

        if "LumenOutput" in out and lumens:
            out["LumenOutput"] = lumens

        # Input load fields: template includes both InputLoad and Lamp1InputLoad; we prefer InputLoad if present
        if input_load:
            if "InputLoad" in out:
                out["InputLoad"] = input_load
            elif "Lamp1InputLoad" in out:
                out["Lamp1InputLoad"] = input_load

        # Unit type field
        if unit_field and unit_field in out:
            out[unit_field] = unit_value

        # VOID flag (must be True if VOID appears anywhere in block)
        if "VOID" in out:
            out["VOID"] = True if void_flag else bool(out["VOID"])  # preserve template default if already True

        out_rows.append(out)

    return pd.DataFrame(out_rows, columns=headers)


# -----------------------------
# UI
# -----------------------------

st.set_page_config(page_title="MF Fixture Schedule -> Template CSV", layout="wide")

st.title("MF Fixture Schedule → TEMPLATE.xlsx CSV Export")

st.markdown(
    """
Upload:
1) **TEMPLATE.xlsx** (must include sheet `tbeFixtureTypeDetails`)
2) Your **MF Fixture Schedule CSV** (the report-style CSV you originally imported)

Then download the output CSV with:
- VOID detection
- UnitType normalization (`ln.ft` vs `each`)
- Location + Protection pulled
- LumenOutput + InputLoad pulled per fixture
- CatalogNo preserved exactly (including *TBC notes)
"""
)

template_file = st.file_uploader("Upload TEMPLATE.xlsx", type=["xlsx"])
schedule_file = st.file_uploader("Upload MF Fixture Schedule CSV", type=["csv"])

colA, colB = st.columns([1, 1])

with colA:
    st.checkbox("Show preview of detected fixture rows", value=False, key="show_preview")

with colB:
    st.checkbox("Show output preview", value=True, key="show_output_preview")

if template_file and schedule_file:
    try:
        headers, default_row = load_template_defaults(template_file.read())
        schedule_df = pd.read_csv(schedule_file, dtype=str)

        if st.session_state.show_preview:
            first_col = schedule_df.columns[0]
            mask = schedule_df[first_col].apply(is_fixture_code)
            st.write("Detected fixture designations (first 50):")
            st.dataframe(schedule_df.loc[mask, [first_col]].head(50), use_container_width=True)

        out_df = transform(schedule_df, headers, default_row)

        if st.session_state.show_output_preview:
            st.subheader("Output preview")
            show_cols = [c for c in ["Type", "Manufacturer", "CatalogNo", "Location", "Protection", "LumenOutput", "InputLoad", "VOID"] if c in out_df.columns]
            st.dataframe(out_df[show_cols].head(50), use_container_width=True)

        # Create CSV bytes
        csv_bytes = out_df.to_csv(index=False).encode("utf-8")

        st.download_button(
            label="Download output CSV",
            data=csv_bytes,
            file_name="fixture_template_export.csv",
            mime="text/csv",
        )

        st.success(f"Done. Export has {out_df.shape[0]} rows and {out_df.shape[1]} columns (template-preserving).")

    except Exception as e:
        st.error(str(e))
else:
    st.info("Upload both files to generate the export.")
